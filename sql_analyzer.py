import sqlparse
import re
import json
import io
import pandas as pd
from typing import Dict, List, Set, Tuple, Optional
from collections import defaultdict

class SQLAnalyzer:
    def __init__(self, dialect: Optional[str] = None):
        """
        Initialize SQL Analyzer
        
        Args:
            dialect: SQL dialect ('postgresql', 'mysql', 'sqlite', or None for auto-detection)
        """
        self.dialect = dialect
        self.temp_table_patterns = [
            r'CREATE\s+TEMP(?:ORARY)?\s+TABLE\s+([^\s\(]+)',
            r'CREATE\s+TEMPORARY\s+TABLE\s+([^\s\(]+)',
            r'CREATE\s+LOCAL\s+TEMPORARY\s+TABLE\s+([^\s\(]+)',
            r'CREATE\s+GLOBAL\s+TEMPORARY\s+TABLE\s+([^\s\(]+)'
        ]
    
    def analyze_sql(self, sql_content: str, filename: str = "unknown") -> Dict:
        """
        Analyze SQL content and extract table information
        
        Args:
            sql_content: Raw SQL content
            filename: Name of the source file
            
        Returns:
            Dictionary containing analysis results
        """
        try:
            # Parse SQL content
            statements = sqlparse.split(sql_content)
            statements = [stmt.strip() for stmt in statements if stmt.strip()]
            
            results = {
                'filename': filename,
                'total_statements': len(statements),
                'from_tables': defaultdict(list),
                'temp_tables': [],
                'ctes': [],
                'dependencies': defaultdict(list),
                'error': None
            }
            
            for i, statement in enumerate(statements, 1):
                try:
                    # Parse individual statement
                    parsed = sqlparse.parse(statement)[0]
                    
                    # Extract FROM tables
                    from_tables = self._extract_from_tables(parsed, i)
                    for table in from_tables:
                        results['from_tables'][table].append(i)
                    
                    # Extract temporary tables
                    temp_tables = self._extract_temp_tables(statement, i)
                    results['temp_tables'].extend(temp_tables)
                    
                    # Extract CTEs
                    ctes = self._extract_ctes(parsed, i)
                    results['ctes'].extend(ctes)
                    
                    # Track dependencies
                    if from_tables:
                        results['dependencies'][i] = from_tables
                        
                except Exception as e:
                    # Continue processing other statements even if one fails
                    print(f"Warning: Error parsing statement {i}: {str(e)}")
                    continue
            
            # Convert defaultdict to regular dict for JSON serialization
            results['from_tables'] = dict(results['from_tables'])
            results['dependencies'] = dict(results['dependencies'])
            
            return results
            
        except Exception as e:
            return {
                'filename': filename,
                'total_statements': 0,
                'from_tables': {},
                'temp_tables': [],
                'ctes': [],
                'dependencies': {},
                'error': str(e)
            }
    
    def _extract_from_tables(self, parsed_stmt, statement_num: int) -> Set[str]:
        """Extract table names from FROM clauses"""
        tables = set()
        statement_str = str(parsed_stmt).upper()
        
        # Use regex patterns to find table names more reliably
        patterns = [
            # FROM table_name
            r'FROM\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)?)',
            # JOIN table_name
            r'JOIN\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)?)',
            # UPDATE table_name
            r'UPDATE\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)?)',
            # INSERT INTO table_name
            r'INTO\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)?)',
        ]
        
        original_stmt = str(parsed_stmt)
        
        for pattern in patterns:
            matches = re.finditer(pattern, original_stmt, re.IGNORECASE)
            for match in matches:
                table_name = self._clean_table_name(match.group(1))
                if table_name and self._is_valid_table_name(table_name):
                    tables.add(table_name)
        
        # Also try the token-based approach for complex cases
        tokens = list(parsed_stmt.flatten())
        for i, token in enumerate(tokens):
            if (token.ttype is sqlparse.tokens.Keyword and 
                token.value.upper() in ['FROM', 'JOIN', 'UPDATE', 'INTO']):
                
                # Look ahead for table names
                for j in range(i + 1, min(i + 10, len(tokens))):
                    next_token = tokens[j]
                    
                    # Skip whitespace and punctuation
                    if next_token.is_whitespace or next_token.value in ['(', ')', ',']:
                        continue
                    
                    # Stop at certain keywords
                    if (next_token.ttype is sqlparse.tokens.Keyword and 
                        next_token.value.upper() in ['WHERE', 'GROUP', 'ORDER', 'HAVING', 
                                                   'LIMIT', 'UNION', 'ON', 'INNER',
                                                   'LEFT', 'RIGHT', 'FULL', 'CROSS', 'AS']):
                        break
                    
                    # Extract table name
                    if next_token.ttype is None or next_token.ttype in [sqlparse.tokens.Name]:
                        table_name = self._clean_table_name(next_token.value)
                        if table_name and self._is_valid_table_name(table_name):
                            tables.add(table_name)
                            break  # Found a table name, move to next FROM/JOIN
        
        # Also check for subqueries and CTEs
        tables.update(self._extract_subquery_tables(parsed_stmt))
        
        return tables
    
    def _extract_subquery_tables(self, parsed_stmt) -> Set[str]:
        """Extract tables from subqueries"""
        tables = set()
        
        def walk_tokens(token):
            if hasattr(token, 'tokens'):
                for sub_token in token.tokens:
                    if isinstance(sub_token, sqlparse.sql.Parenthesis):
                        # Parse subquery
                        subquery_content = str(sub_token)[1:-1]  # Remove parentheses
                        try:
                            sub_parsed = sqlparse.parse(subquery_content)
                            if sub_parsed:
                                sub_tables = self._extract_from_tables(sub_parsed[0], 0)
                                tables.update(sub_tables)
                        except:
                            pass
                    else:
                        walk_tokens(sub_token)
        
        walk_tokens(parsed_stmt)
        return tables
    
    def _extract_temp_tables(self, statement: str, statement_num: int) -> List[Dict]:
        """Extract temporary table definitions"""
        temp_tables = []
        
        for pattern in self.temp_table_patterns:
            matches = re.finditer(pattern, statement, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                table_name = self._clean_table_name(match.group(1))
                if table_name:
                    temp_tables.append({
                        'name': table_name,
                        'type': 'TEMPORARY TABLE',
                        'statement_num': statement_num,
                        'definition': statement.strip()
                    })
        
        return temp_tables
    
    def _extract_ctes(self, parsed_stmt, statement_num: int) -> List[Dict]:
        """Extract Common Table Expressions (CTEs)"""
        ctes = []
        statement_str = str(parsed_stmt)
        
        # Look for WITH clauses
        with_pattern = r'WITH\s+(\w+)\s+AS\s*\('
        matches = re.finditer(with_pattern, statement_str, re.IGNORECASE | re.MULTILINE)
        
        for match in matches:
            cte_name = match.group(1)
            
            # Try to extract the full CTE definition
            start_pos = match.start()
            paren_count = 0
            definition_end = start_pos
            
            for i, char in enumerate(statement_str[match.end():], match.end()):
                if char == '(':
                    paren_count += 1
                elif char == ')':
                    paren_count -= 1
                    if paren_count == 0:
                        definition_end = i + 1
                        break
            
            definition = statement_str[start_pos:definition_end]
            
            ctes.append({
                'name': cte_name,
                'type': 'CTE',
                'statement_num': statement_num,
                'definition': definition
            })
        
        return ctes
    
    def _clean_table_name(self, name: str) -> str:
        """Clean and normalize table name"""
        if not name:
            return ""
        
        # Remove quotes and brackets
        name = re.sub(r'^["`\[\]]+|["`\[\]]+$', '', name.strip())
        
        # Remove schema/database prefix (keep only table name)
        if '.' in name:
            parts = name.split('.')
            name = parts[-1]  # Get the last part (table name)
        
        # Remove alias keywords
        name = re.sub(r'\s+AS\s+.*$', '', name, flags=re.IGNORECASE)
        
        return name.strip()
    
    def _is_valid_table_name(self, name: str) -> bool:
        """Check if the extracted name is likely a valid table name"""
        if not name:
            return False
        
        # Skip tables that start with "sessionid" (temporary tables)
        if name.lower().startswith('sessionid'):
            return False
        
        # Skip common database names
        database_names = {
            'cstonedb3', 'mysql', 'postgres', 'oracle', 'sqlserver', 'sqlite',
            'information_schema', 'performance_schema', 'sys', 'master', 'tempdb', 'model', 'msdb'
        }
        
        if name.lower() in database_names:
            return False
        
        # Skip SQL keywords and common non-table tokens
        sql_keywords = {
            'SELECT', 'FROM', 'WHERE', 'GROUP', 'ORDER', 'HAVING', 'LIMIT',
            'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER',
            'JOIN', 'INNER', 'LEFT', 'RIGHT', 'FULL', 'CROSS', 'ON',
            'AND', 'OR', 'NOT', 'IN', 'EXISTS', 'BETWEEN', 'LIKE',
            'IS', 'NULL', 'TRUE', 'FALSE', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END',
            'UNION', 'INTERSECT', 'EXCEPT', 'ALL', 'DISTINCT', 'AS',
            'ASC', 'DESC', 'BY', 'INTO', 'VALUES', 'SET', 'TABLE', 'TEMP',
            'TEMPORARY', 'WITH', 'IF', 'NOT', 'EXISTS'
        }
        
        if name.upper() in sql_keywords:
            return False
        
        # More flexible pattern to allow various table name formats
        if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_#$]*$', name):
            return False
        
        # Skip very short names that are likely not table names
        if len(name) < 2:
            return False
        
        return True
    
    def export_to_csv(self, results: Dict) -> str:
        """Export analysis results to CSV format"""
        output = io.StringIO()
        
        # Write header
        output.write("Type,Name,Statement_Number,Usage_Count,Definition\n")
        
        # Write FROM tables
        for table_name, occurrences in results['from_tables'].items():
            output.write(f"FROM_TABLE,{table_name},{';'.join(map(str, occurrences))},{len(occurrences)},\n")
        
        # Write temporary tables
        for temp_table in results['temp_tables']:
            definition = temp_table['definition'].replace('\n', ' ').replace(',', ';')
            output.write(f"TEMP_TABLE,{temp_table['name']},{temp_table['statement_num']},1,\"{definition}\"\n")
        
        # Write CTEs
        for cte in results['ctes']:
            definition = cte['definition'].replace('\n', ' ').replace(',', ';')
            output.write(f"CTE,{cte['name']},{cte['statement_num']},1,\"{definition}\"\n")
        
        return output.getvalue()
    
    def export_to_json(self, results: Dict) -> str:
        """Export analysis results to JSON format"""
        return json.dumps(results, indent=2, ensure_ascii=False)
    
    def export_to_excel(self, results: Dict) -> bytes:
        """Export analysis results to Excel format"""
        output = io.BytesIO()
        
        # Create workbook and worksheets
        from openpyxl import Workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(['Metric', 'Count'])
        ws_summary.append(['Total Statements', results['total_statements']])
        ws_summary.append(['Unique Tables', len(results['from_tables'])])
        ws_summary.append(['Temporary Tables', len(results['temp_tables'])])
        ws_summary.append(['CTEs', len(results['ctes'])])
        
        # FROM Tables Sheet
        if results['from_tables']:
            ws_tables = wb.create_sheet("FROM Tables")
            ws_tables.append(['Table Name', 'Occurrences', 'Statements', 'First Appearance'])
            
            for table_name, occurrences in results['from_tables'].items():
                ws_tables.append([
                    table_name,
                    len(occurrences),
                    ', '.join(map(str, occurrences)),
                    f"Statement {min(occurrences)}"
                ])
        
        # Temporary Tables Sheet
        if results['temp_tables'] or results['ctes']:
            ws_temp = wb.create_sheet("Temporary Tables & CTEs")
            ws_temp.append(['Name', 'Type', 'Statement Number', 'Definition'])
            
            # Add temporary tables
            for temp_table in results['temp_tables']:
                ws_temp.append([
                    temp_table['name'],
                    temp_table['type'],
                    temp_table['statement_num'],
                    temp_table['definition']
                ])
            
            # Add CTEs
            for cte in results['ctes']:
                ws_temp.append([
                    cte['name'],
                    'CTE',
                    cte['statement_num'],
                    cte['definition']
                ])
        
        # Dependencies Sheet
        if results['dependencies']:
            ws_deps = wb.create_sheet("Dependencies")
            ws_deps.append(['Statement Number', 'Tables Used'])
            
            for stmt_num, deps in results['dependencies'].items():
                if deps:
                    ws_deps.append([stmt_num, ', '.join(deps)])
        
        # Save to BytesIO
        wb.save(output)
        return output.getvalue()
